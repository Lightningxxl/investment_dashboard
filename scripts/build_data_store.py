#!/usr/bin/env python3
"""Build structured dashboard data from the recommended-target ZIP package."""

from __future__ import annotations

import json
import math
import re
import shutil
import sqlite3
import sys
import zipfile
from datetime import date, datetime
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from statistics import mean, stdev

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DEFAULT_ZIP = Path.home() / "Downloads" / "中国50研选池推荐标的整理.zip"
SOURCE_JSON = DATA_DIR / "source_china50_fund_pool.json"
SOURCE_JSON_FALLBACK = ROOT / "china50_fund_pool_data.json"
SOURCE_CSV_FALLBACK = ROOT / "china50_fund_pool_data.csv"
OUT_SQLITE = DATA_DIR / "investment_dashboard.sqlite"
OUT_JSON = DATA_DIR / "dashboard_data.json"
OUT_JS = DATA_DIR / "dashboard_data.js"
OUT_CSV = DATA_DIR / "china50_fund_pool.csv"

MANUAL_MATCHES = {
    "诚奇金选500指数增强1期": "诚奇金选500指数增强1号B[1]",
    "聚鸣金选高山8号私募证券投资基金1期": "聚鸣金选高山8号B[1]",
    "衍复金选沪深300指数增强一号私募证券投资": "衍复金选沪深300指数增强一号B[1]",
}

PERCENT_METRICS = {
    "年化收益率",
    "阿尔法",
    "胜率",
    "年化波动",
    "最大回撤",
    "下行风险",
}

CHINESE_NUMERAL_REPLACEMENTS = {
    "一": "1",
    "二": "2",
    "三": "3",
    "四": "4",
    "五": "5",
    "六": "6",
    "七": "7",
    "八": "8",
    "九": "9",
    "十": "10",
    "壹": "1",
    "贰": "2",
    "叁": "3",
}


def decode_zip_name(name: str) -> str:
    try:
        return name.encode("cp437").decode("utf-8")
    except Exception:
        return name


def parse_date(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass
    return None


def parse_num(value, *, percent: bool | None = None):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if not math.isfinite(float(value)):
            return None
        return float(value) * 100 if percent is True else float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "--", "nan", "None"}:
        return None
    has_percent = "%" in text
    text = text.replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return None
    if percent is False:
        return number
    return number if has_percent or percent is True else number


def parse_days(value):
    if value is None or value == "":
        return None
    match = re.search(r"(-?\d+(?:\.\d+)?)", str(value))
    return int(float(match.group(1))) if match else None


def clean_workbook_rows(ws):
    try:
        ws.reset_dimensions()
    except Exception:
        pass
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if any(value not in (None, "") for value in values):
            rows.append(values)
    return rows


def parse_filename(path: str) -> tuple[str, str, str | None]:
    base = Path(path).name
    report_date = None
    match = re.search(r"业绩表现\((\d{4}-\d{2}-\d{2})\)\.xlsx$", base)
    if match:
        report_date = match.group(1)
    base = re.sub(r"-业绩表现\(\d{4}-\d{2}-\d{2}\)\.xlsx$", "", base)
    code_match = re.match(r"(.+)-([A-Z0-9]+)$", base)
    if code_match:
        return code_match.group(1), code_match.group(2), report_date
    return base, "", report_date


def parse_series(rows, fields, *, start_row=1):
    series = []
    for row in rows[start_row:]:
        row = list(row) + [None] * max(0, len(fields) - len(row))
        row_date = parse_date(row[0])
        if not row_date:
            continue
        item = {"date": row_date}
        for index, (field, kind) in enumerate(fields, start=1):
            value = row[index] if index < len(row) else None
            if kind == "pct":
                item[field] = parse_num(value, percent=True)
            elif kind == "num":
                item[field] = parse_num(value, percent=False)
            else:
                item[field] = "" if value is None else str(value)
        series.append(item)
    series.sort(key=lambda row: row["date"])
    return series


def parse_interval_returns(rows):
    if not rows:
        return {}
    result = {}
    headers = rows[0]
    benchmark_name = headers[2] if len(headers) > 2 else None
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        result[str(row[0])] = {
            "product_return_pct": parse_num(row[1] if len(row) > 1 else None, percent=True),
            "benchmark_return_pct": parse_num(row[2] if len(row) > 2 else None, percent=True),
            "excess_return_pct": parse_num(row[3] if len(row) > 3 else None, percent=True),
        }
    return result, benchmark_name


def parse_metric_table(rows):
    if not rows:
        return {}
    horizons = [str(value) for value in rows[0][1:] if value not in (None, "")]
    result = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        metric = str(row[0])
        result[metric] = {}
        for horizon, value in zip(horizons, row[1:]):
            if metric == "回撤修复":
                parsed = parse_days(value)
            elif metric in PERCENT_METRICS:
                parsed = parse_num(value, percent=True)
            else:
                parsed = parse_num(value, percent=False)
            result[metric][horizon] = parsed
    return result


def first_metric(metrics, metric, horizons):
    row = metrics.get(metric, {})
    for horizon in horizons:
        value = row.get(horizon)
        if value is not None:
            return value
    return None


def annualized_from_nav(nav_series):
    if len(nav_series) < 2:
        return None
    start = nav_series[0]
    end = nav_series[-1]
    start_nav = start.get("adjusted_nav") or start.get("acc_nav") or start.get("unit_nav")
    end_nav = end.get("adjusted_nav") or end.get("acc_nav") or end.get("unit_nav")
    if not start_nav or not end_nav or start_nav <= 0:
        return None
    start_date = datetime.fromisoformat(start["date"]).date()
    end_date = datetime.fromisoformat(end["date"]).date()
    years = max((end_date - start_date).days / 365.25, 1 / 365.25)
    return ((end_nav / start_nav) ** (1 / years) - 1) * 100


def periodic_returns_from_nav(nav_series):
    returns = []
    for prev, current in zip(nav_series, nav_series[1:]):
        prev_nav = prev.get("adjusted_nav") or prev.get("acc_nav") or prev.get("unit_nav")
        current_nav = current.get("adjusted_nav") or current.get("acc_nav") or current.get("unit_nav")
        if prev_nav and current_nav and prev_nav > 0:
            returns.append({"date": current["date"], "return_pct": (current_nav / prev_nav - 1) * 100})
    return returns


def simple_corr(a, b):
    if len(a) < 4 or len(b) < 4 or len(a) != len(b):
        return None
    try:
        am = mean(a)
        bm = mean(b)
        den_a = math.sqrt(sum((value - am) ** 2 for value in a))
        den_b = math.sqrt(sum((value - bm) ** 2 for value in b))
        if not den_a or not den_b:
            return None
        return max(-1, min(1, sum((x - am) * (y - bm) for x, y in zip(a, b)) / (den_a * den_b)))
    except Exception:
        return None


def benchmark_correlation(return_series):
    if len(return_series) < 5:
        return None
    product = []
    benchmark = []
    for prev, current in zip(return_series, return_series[1:]):
        if prev.get("product_return_pct") is None or current.get("product_return_pct") is None:
            continue
        if prev.get("benchmark_return_pct") is None or current.get("benchmark_return_pct") is None:
            continue
        product.append(current["product_return_pct"] - prev["product_return_pct"])
        benchmark.append(current["benchmark_return_pct"] - prev["benchmark_return_pct"])
    return simple_corr(product, benchmark)


def max_drawdown_from_nav(nav_series):
    peak = None
    max_dd = 0
    for row in nav_series:
        nav = row.get("adjusted_nav") or row.get("acc_nav") or row.get("unit_nav")
        if not nav:
            continue
        peak = nav if peak is None else max(peak, nav)
        max_dd = min(max_dd, nav / peak - 1)
    return abs(max_dd * 100)


def nav_years(nav_series):
    if len(nav_series) < 2:
        return 0
    start = datetime.fromisoformat(nav_series[0]["date"]).date()
    end = datetime.fromisoformat(nav_series[-1]["date"]).date()
    return max(0, (end - start).days / 365.25)


def parse_workbook(zip_file, raw_name):
    display_name = decode_zip_name(raw_name)
    product_name, product_code, report_date = parse_filename(display_name)
    category_path = "/".join(display_name.split("/")[1:-1])
    workbook = openpyxl.load_workbook(BytesIO(zip_file.read(raw_name)), data_only=True, read_only=True)
    sheet_rows = {
        sheet_name: clean_workbook_rows(workbook[sheet_name])
        for sheet_name in workbook.sheetnames
    }

    nav_series = parse_series(
        sheet_rows.get("净值走势", []),
        [
            ("unit_nav", "num"),
            ("acc_nav", "num"),
            ("acc_nav_change_pct", "pct"),
            ("adjusted_nav", "num"),
            ("adjusted_nav_change_pct", "pct"),
            ("dividend", "num"),
        ],
    )
    return_series = parse_series(
        sheet_rows.get("收益走势", []),
        [
            ("product_return_pct", "pct"),
            ("benchmark_return_pct", "pct"),
            ("excess_return_pct", "pct"),
        ],
    )
    drawdown_series = parse_series(
        sheet_rows.get("动态回撤", []),
        [
            ("product_drawdown_pct", "pct"),
            ("benchmark_drawdown_pct", "pct"),
            ("excess_drawdown_pct", "pct"),
        ],
    )
    interval_returns, interval_benchmark = parse_interval_returns(sheet_rows.get("区间收益", []))
    return_metrics = parse_metric_table(sheet_rows.get("收益指标", []))
    risk_metrics = parse_metric_table(sheet_rows.get("风险指标", []))
    scale_rows = sheet_rows.get("产品规模", [])
    scale_series = parse_series(scale_rows, [("scale_wan", "num")], start_row=2)

    benchmark_name = None
    returns_header = sheet_rows.get("收益走势", [[]])[0] if sheet_rows.get("收益走势") else []
    if len(returns_header) > 2:
        benchmark_name = returns_header[2]
    benchmark_name = benchmark_name or interval_benchmark

    periodic_returns = periodic_returns_from_nav(nav_series)
    inception_return = first_metric(return_metrics, "年化收益率", ["成立以来", "近一年"])
    inception_vol = first_metric(risk_metrics, "年化波动", ["成立以来", "近一年"])
    inception_dd = first_metric(risk_metrics, "最大回撤", ["成立以来", "近一年"])
    inception_beta = first_metric(risk_metrics, "贝塔", ["成立以来", "近一年"])
    inception_sharpe = first_metric(return_metrics, "夏普比率", ["成立以来", "近一年"])
    computed_dd = max_drawdown_from_nav(nav_series)

    summary = {
        "nav_start_date": nav_series[0]["date"] if nav_series else None,
        "nav_end_date": nav_series[-1]["date"] if nav_series else None,
        "nav_points": len(nav_series),
        "nav_years": nav_years(nav_series),
        "return_points": len(return_series),
        "drawdown_points": len(drawdown_series),
        "scale_points": len(scale_series),
        "latest_scale_date": scale_series[-1]["date"] if scale_series else None,
        "latest_scale_wan": scale_series[-1]["scale_wan"] if scale_series else None,
        "latest_scale_yi": (scale_series[-1]["scale_wan"] / 10000) if scale_series else None,
        "benchmark_name": benchmark_name,
        "annual_return_inception_pct": inception_return if inception_return is not None else annualized_from_nav(nav_series),
        "annual_return_1y_pct": first_metric(return_metrics, "年化收益率", ["近一年"]),
        "annual_vol_inception_pct": inception_vol,
        "annual_vol_1y_pct": first_metric(risk_metrics, "年化波动", ["近一年"]),
        "max_drawdown_inception_abs_pct": abs(inception_dd) if inception_dd is not None else computed_dd,
        "max_drawdown_1y_abs_pct": abs(first_metric(risk_metrics, "最大回撤", ["近一年"]) or 0) or None,
        "beta_inception": inception_beta,
        "beta_1y": first_metric(risk_metrics, "贝塔", ["近一年"]),
        "sharpe_inception": inception_sharpe,
        "sharpe_1y": first_metric(return_metrics, "夏普比率", ["近一年"]),
        "benchmark_correlation": benchmark_correlation(return_series),
    }

    return {
        "source_file": display_name,
        "category_path": category_path,
        "category": category_path.split("/")[0] if category_path else "",
        "product_name": product_name,
        "product_code": product_code,
        "report_date": report_date,
        "summary": summary,
        "series": {
            "nav": nav_series,
            "periodic_returns": periodic_returns,
            "returns": return_series,
            "drawdowns": drawdown_series,
            "scale": scale_series,
        },
        "tables": {
            "interval_returns": interval_returns,
            "return_metrics": return_metrics,
            "risk_metrics": risk_metrics,
        },
    }


def replace_chinese_number_words(text: str) -> str:
    result = text
    for source, target in CHINESE_NUMERAL_REPLACEMENTS.items():
        result = result.replace(f"{source}号", f"{target}号").replace(f"{source}期", f"{target}期")
    return result


def normalize_name(value: str) -> str:
    text = replace_chinese_number_words(str(value or ""))
    text = re.sub(r"\[[123]\]", "", text).lower()
    replacements = [
        ("私募证券投资基金", ""),
        ("证券投资基金", ""),
        ("私募投资基金", ""),
        ("私募基金", ""),
        ("基金", ""),
        ("b类", "b"),
        ("a类", "a"),
        ("份额", ""),
        ("金选专享", "金选"),
    ]
    for source, target in replacements:
        text = text.replace(source, target)
    text = re.sub(r"[\s·・_\-—:：,，()（）/\\]+", "", text)
    return text


def extract_family_tags(value: str) -> set[str]:
    text = normalize_name(value)
    tags = set()
    for tag in ["a500", "1000", "500", "300", "2000", "全指", "量化对冲", "量化选股", "cta", "多策略", "港股通"]:
        if tag in text:
            tags.add(tag)
    if "a500" in tags:
        tags.discard("500")
    return tags


def extract_numbers(value: str) -> set[str]:
    return set(re.findall(r"(?<![a-z])(?:a?\d{2,4}|\d+)(?![a-z])", normalize_name(value)))


def match_score(target_name: str, record: dict) -> float:
    target_norm = normalize_name(target_name)
    product_norm = normalize_name(record.get("product_name", ""))
    score = SequenceMatcher(None, target_norm, product_norm).ratio()
    if target_norm in product_norm or product_norm in target_norm:
        score = max(score, 0.96 if min(len(target_norm), len(product_norm)) > 5 else 0.8)
    manager = record.get("manager_name_guess") or ""
    if manager and manager in target_name:
        score += 0.02

    critical = {"a500", "1000", "500", "300", "2000", "全指", "港股通"}
    if (extract_family_tags(target_name) & critical) != (extract_family_tags(record.get("product_name", "")) & critical):
        score -= 0.18

    target_numbers = extract_numbers(target_name)
    product_numbers = extract_numbers(record.get("product_name", ""))
    if target_numbers and product_numbers and not target_numbers.intersection(product_numbers):
        score -= 0.12
    return score


def source_json_path() -> Path:
    return SOURCE_JSON if SOURCE_JSON.exists() else SOURCE_JSON_FALLBACK


def load_base_source():
    return json.loads(source_json_path().read_text(encoding="utf-8"))


def attach_matches(targets, source=None):
    source = source or load_base_source()
    products = [
        record
        for record in source.get("records", [])
        if record.get("row_type") == "product" and record.get("product_name") and record.get("product_name") != "-"
    ]
    by_name = {record["product_name"]: record for record in products}

    for target in targets:
        manual_name = MANUAL_MATCHES.get(target["product_name"])
        if manual_name and manual_name in by_name:
            best_record = by_name[manual_name]
            best_score = 1.0
            confidence = "high"
            method = "manual"
        else:
            candidates = sorted(
                ((match_score(target["product_name"], record), record) for record in products),
                key=lambda item: item[0],
                reverse=True,
            )
            best_score, best_record = candidates[0]
            if best_score >= 0.90:
                confidence = "high"
            elif best_score >= 0.78:
                confidence = "review"
            else:
                confidence = "unmatched"
            method = "name_similarity"

        target["match"] = {
            "confidence": confidence,
            "method": method,
            "score": round(best_score, 4),
            "pdf_product_name": None if confidence == "unmatched" else best_record.get("product_name"),
            "pdf_manager_label": None if confidence == "unmatched" else best_record.get("manager_fund_manager_strategy_label"),
            "pdf_manager_name": None if confidence == "unmatched" else best_record.get("manager_name_guess"),
            "pdf_section": None if confidence == "unmatched" else best_record.get("section"),
            "pdf_subsection": None if confidence == "unmatched" else best_record.get("subsection"),
            "pdf_source_page": None if confidence == "unmatched" else best_record.get("source_page"),
            "nearest_pdf_product_name": best_record.get("product_name"),
            "nearest_pdf_manager_label": best_record.get("manager_fund_manager_strategy_label"),
        }
    return targets


def parse_recommended_targets(zip_path: Path, source):
    with zipfile.ZipFile(zip_path) as zip_file:
        workbook_names = [
            name
            for name in zip_file.namelist()
            if name.endswith(".xlsx") and not name.startswith("__MACOSX/") and "/._" not in name
        ]
        targets = [parse_workbook(zip_file, name) for name in workbook_names]

    targets.sort(key=lambda item: (item["category_path"], item["product_name"]))
    attach_matches(targets, source)
    return targets


def recommended_metadata(zip_path: Path, targets):
    latest_scale_sum = sum(target["summary"].get("latest_scale_wan") or 0 for target in targets)
    counts = {}
    for target in targets:
        confidence = target["match"]["confidence"]
        counts[confidence] = counts.get(confidence, 0) + 1
    return {
        "source_zip": zip_path.name,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_report_date": "2026-06-23",
        "target_count": len(targets),
        "high_confidence_match_count": counts.get("high", 0),
        "review_match_count": counts.get("review", 0),
        "unmatched_count": counts.get("unmatched", 0),
        "latest_scale_total_wan": latest_scale_sum,
        "latest_scale_total_yi": latest_scale_sum / 10000,
        "notes": [
            "Excel sheets use incorrect worksheet dimension metadata; parser resets worksheet dimensions before reading.",
            "High-confidence matches are merged into the dashboard model. Review/unmatched targets stay in this dataset for manual validation.",
            "Percentage values are stored in percentage points, e.g. 12.3 means 12.3%.",
        ],
    }


def json_text(value):
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def create_schema(conn):
    conn.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE metadata (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL
        );

        CREATE TABLE pdf_records (
          id INTEGER PRIMARY KEY,
          row_type TEXT,
          source_page INTEGER,
          section TEXT,
          subsection TEXT,
          product_name TEXT,
          manager_label TEXT,
          manager_name TEXT,
          is_subscription_paused INTEGER,
          record_json TEXT NOT NULL
        );

        CREATE TABLE recommended_targets (
          id INTEGER PRIMARY KEY,
          product_name TEXT NOT NULL,
          product_code TEXT,
          category_path TEXT,
          category TEXT,
          source_file TEXT,
          report_date TEXT,
          summary_json TEXT NOT NULL,
          target_json TEXT NOT NULL
        );

        CREATE TABLE fund_matches (
          target_id INTEGER PRIMARY KEY,
          confidence TEXT NOT NULL,
          method TEXT,
          score REAL,
          pdf_product_name TEXT,
          pdf_manager_label TEXT,
          pdf_manager_name TEXT,
          pdf_section TEXT,
          pdf_subsection TEXT,
          pdf_source_page INTEGER,
          nearest_pdf_product_name TEXT,
          nearest_pdf_manager_label TEXT,
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE nav_series (
          target_id INTEGER NOT NULL,
          date TEXT NOT NULL,
          unit_nav REAL,
          acc_nav REAL,
          acc_nav_change_pct REAL,
          adjusted_nav REAL,
          adjusted_nav_change_pct REAL,
          dividend REAL,
          PRIMARY KEY(target_id, date),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE periodic_returns (
          target_id INTEGER NOT NULL,
          date TEXT NOT NULL,
          return_pct REAL,
          PRIMARY KEY(target_id, date),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE return_series (
          target_id INTEGER NOT NULL,
          date TEXT NOT NULL,
          product_return_pct REAL,
          benchmark_return_pct REAL,
          excess_return_pct REAL,
          PRIMARY KEY(target_id, date),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE drawdown_series (
          target_id INTEGER NOT NULL,
          date TEXT NOT NULL,
          product_drawdown_pct REAL,
          benchmark_drawdown_pct REAL,
          excess_drawdown_pct REAL,
          PRIMARY KEY(target_id, date),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE aum_series (
          target_id INTEGER NOT NULL,
          date TEXT NOT NULL,
          scale_wan REAL,
          PRIMARY KEY(target_id, date),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE interval_returns (
          target_id INTEGER NOT NULL,
          horizon TEXT NOT NULL,
          product_return_pct REAL,
          benchmark_return_pct REAL,
          excess_return_pct REAL,
          PRIMARY KEY(target_id, horizon),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE return_metrics (
          target_id INTEGER NOT NULL,
          metric TEXT NOT NULL,
          horizon TEXT NOT NULL,
          value REAL,
          PRIMARY KEY(target_id, metric, horizon),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE TABLE risk_metrics (
          target_id INTEGER NOT NULL,
          metric TEXT NOT NULL,
          horizon TEXT NOT NULL,
          value REAL,
          PRIMARY KEY(target_id, metric, horizon),
          FOREIGN KEY(target_id) REFERENCES recommended_targets(id) ON DELETE CASCADE
        );

        CREATE INDEX idx_pdf_records_product_name ON pdf_records(product_name);
        CREATE INDEX idx_pdf_records_section ON pdf_records(section);
        CREATE INDEX idx_recommended_targets_product_name ON recommended_targets(product_name);
        CREATE INDEX idx_fund_matches_confidence ON fund_matches(confidence);
        """
    )


def write_sqlite(source, recommended_payload):
    if OUT_SQLITE.exists():
        OUT_SQLITE.unlink()
    conn = sqlite3.connect(OUT_SQLITE)
    create_schema(conn)
    meta = {
        "dashboard_generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_pdf": source.get("metadata", {}).get("source_pdf", ""),
        "source_report_date": source.get("metadata", {}).get("report_date", ""),
        "pdf_record_count": len(source.get("records", [])),
        "recommended_target_count": recommended_payload["metadata"]["target_count"],
        "high_confidence_match_count": recommended_payload["metadata"]["high_confidence_match_count"],
    }
    conn.executemany(
        "INSERT INTO metadata(key, value) VALUES (?, ?)",
        [(key, json_text(value) if isinstance(value, (dict, list)) else str(value)) for key, value in meta.items()],
    )
    for index, record in enumerate(source.get("records", []), start=1):
        conn.execute(
            """
            INSERT INTO pdf_records
              (id, row_type, source_page, section, subsection, product_name, manager_label,
               manager_name, is_subscription_paused, record_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                index,
                record.get("row_type"),
                record.get("source_page"),
                record.get("section"),
                record.get("subsection"),
                record.get("product_name"),
                record.get("manager_fund_manager_strategy_label"),
                record.get("manager_name_guess"),
                1 if record.get("is_subscription_paused") else 0,
                json_text(record),
            ),
        )

    for target_id, target in enumerate(recommended_payload["targets"], start=1):
        conn.execute(
            """
            INSERT INTO recommended_targets
              (id, product_name, product_code, category_path, category, source_file,
               report_date, summary_json, target_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                target_id,
                target["product_name"],
                target.get("product_code"),
                target.get("category_path"),
                target.get("category"),
                target.get("source_file"),
                target.get("report_date"),
                json_text(target.get("summary", {})),
                json_text(target),
            ),
        )
        match = target.get("match", {})
        conn.execute(
            """
            INSERT INTO fund_matches
              (target_id, confidence, method, score, pdf_product_name, pdf_manager_label,
               pdf_manager_name, pdf_section, pdf_subsection, pdf_source_page,
               nearest_pdf_product_name, nearest_pdf_manager_label)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                target_id,
                match.get("confidence"),
                match.get("method"),
                match.get("score"),
                match.get("pdf_product_name"),
                match.get("pdf_manager_label"),
                match.get("pdf_manager_name"),
                match.get("pdf_section"),
                match.get("pdf_subsection"),
                match.get("pdf_source_page"),
                match.get("nearest_pdf_product_name"),
                match.get("nearest_pdf_manager_label"),
            ),
        )
        for row in target.get("series", {}).get("nav", []):
            conn.execute(
                """
                INSERT INTO nav_series
                  (target_id, date, unit_nav, acc_nav, acc_nav_change_pct,
                   adjusted_nav, adjusted_nav_change_pct, dividend)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    target_id,
                    row.get("date"),
                    row.get("unit_nav"),
                    row.get("acc_nav"),
                    row.get("acc_nav_change_pct"),
                    row.get("adjusted_nav"),
                    row.get("adjusted_nav_change_pct"),
                    row.get("dividend"),
                ),
            )
        for row in target.get("series", {}).get("periodic_returns", []):
            conn.execute(
                "INSERT INTO periodic_returns(target_id, date, return_pct) VALUES (?, ?, ?)",
                (target_id, row.get("date"), row.get("return_pct")),
            )
        for row in target.get("series", {}).get("returns", []):
            conn.execute(
                """
                INSERT INTO return_series
                  (target_id, date, product_return_pct, benchmark_return_pct, excess_return_pct)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    target_id,
                    row.get("date"),
                    row.get("product_return_pct"),
                    row.get("benchmark_return_pct"),
                    row.get("excess_return_pct"),
                ),
            )
        for row in target.get("series", {}).get("drawdowns", []):
            conn.execute(
                """
                INSERT INTO drawdown_series
                  (target_id, date, product_drawdown_pct, benchmark_drawdown_pct, excess_drawdown_pct)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    target_id,
                    row.get("date"),
                    row.get("product_drawdown_pct"),
                    row.get("benchmark_drawdown_pct"),
                    row.get("excess_drawdown_pct"),
                ),
            )
        for row in target.get("series", {}).get("scale", []):
            conn.execute(
                "INSERT INTO aum_series(target_id, date, scale_wan) VALUES (?, ?, ?)",
                (target_id, row.get("date"), row.get("scale_wan")),
            )
        for horizon, row in target.get("tables", {}).get("interval_returns", {}).items():
            conn.execute(
                """
                INSERT INTO interval_returns
                  (target_id, horizon, product_return_pct, benchmark_return_pct, excess_return_pct)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    target_id,
                    horizon,
                    row.get("product_return_pct"),
                    row.get("benchmark_return_pct"),
                    row.get("excess_return_pct"),
                ),
            )
        for metric, values in target.get("tables", {}).get("return_metrics", {}).items():
            for horizon, value in values.items():
                conn.execute(
                    "INSERT INTO return_metrics(target_id, metric, horizon, value) VALUES (?, ?, ?, ?)",
                    (target_id, metric, horizon, value),
                )
        for metric, values in target.get("tables", {}).get("risk_metrics", {}).items():
            for horizon, value in values.items():
                conn.execute(
                    "INSERT INTO risk_metrics(target_id, metric, horizon, value) VALUES (?, ?, ?, ?)",
                    (target_id, metric, horizon, value),
                )
    conn.commit()
    conn.close()


def write_frontend_files(source, recommended_payload, zip_path):
    payload = {
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "sqlite_file": OUT_SQLITE.name,
            "source_pdf": source.get("metadata", {}).get("source_pdf"),
            "pdf_report_date": source.get("metadata", {}).get("report_date"),
            "recommended_source_zip": zip_path.name,
            "pdf_record_count": len(source.get("records", [])),
            "pdf_product_count": source.get("metadata", {}).get("product_count"),
            "recommended_target_count": recommended_payload["metadata"]["target_count"],
            "high_confidence_match_count": recommended_payload["metadata"]["high_confidence_match_count"],
            "review_match_count": recommended_payload["metadata"]["review_match_count"],
            "unmatched_count": recommended_payload["metadata"]["unmatched_count"],
        },
        "basePool": source,
        "recommendedTargets": {
            "metadata": recommended_payload["metadata"],
            "targets": recommended_payload["targets"],
        },
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_JS.write_text(
        "window.DASHBOARD_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    return payload


def copy_source_files():
    if SOURCE_JSON_FALLBACK.exists() and not SOURCE_JSON.exists():
        shutil.copyfile(SOURCE_JSON_FALLBACK, SOURCE_JSON)
    if SOURCE_CSV_FALLBACK.exists() and not OUT_CSV.exists():
        shutil.copyfile(SOURCE_CSV_FALLBACK, OUT_CSV)


def build(zip_path: Path):
    DATA_DIR.mkdir(exist_ok=True)
    copy_source_files()
    source = load_base_source()
    targets = parse_recommended_targets(zip_path, source)
    recommended_payload = {
        "metadata": recommended_metadata(zip_path, targets),
        "targets": targets,
    }
    write_sqlite(source, recommended_payload)
    payload = write_frontend_files(source, recommended_payload, zip_path)
    return payload


def main():
    zip_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else DEFAULT_ZIP
    if not zip_path.exists():
        raise SystemExit(f"ZIP not found: {zip_path}")
    payload = build(zip_path)
    meta = payload["recommendedTargets"]["metadata"]
    print(f"pdf records: {payload['metadata']['pdf_record_count']}")
    print(f"targets: {meta['target_count']}")
    print(f"high confidence matches: {meta['high_confidence_match_count']}")
    print(f"review matches: {meta['review_match_count']}")
    print(f"unmatched: {meta['unmatched_count']}")
    print(f"latest scale total: {meta['latest_scale_total_yi']:.2f} 亿")
    print(f"wrote: {OUT_SQLITE}")
    print(f"wrote: {OUT_JSON}")
    print(f"wrote: {OUT_JS}")


if __name__ == "__main__":
    main()
